from unittest import result
import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook
import lxml
day="2/22/2020"
url=rf"https://www.yallakora.com/match-center/%d9%85%d8%b1%d9%83%d8%b2-%d8%a7%d9%84%d9%85%d8%a8%d8%a7%d8%b1%d9%8a%d8%a7%d8%aa?date={day}#"

page=requests.get(url=url)

# 
#make all programe in this function
def main(page):
    src=page.content # return the content as byte code can't be readable so make [paring]
    # print(src) -> 

    # parsing by beautifulsoup (bytecode,parser)
    parsing=BeautifulSoup(src,"lxml")
    # print(parsing) -> retun HTML readable
    find_all_content_list=parsing.find_all("div",{"class":"matchCard"}) # return a list[] with all elements
    # print(find_all_content_list) -> []
    for mc in range(len(find_all_content_list)):
        #extract the info from 1st matchcard
        matchcard=find_all_content_list[mc]
        # title for the first match card
        champion_title=matchcard.find("h2").text.strip()
        # mathches states
        matchstate=matchcard.find_all("div",{"class":"date"})
        # Teams A
        teamA=matchcard.find_all("div",{"class":"teams teamA"})
        # Teams B
        teamB=matchcard.find_all("div",{"class":"teams teamB"})
        # scores
        scores=matchcard.find_all("span",{"class":"score"})
        scoresA=[]
        scoresB=[]
        for i in range(len(scores)):
            if i==0 or i%2==0:
                scoresA.append(scores[i])

            else:
                scoresB.append(scores[i])



        # print(scores) #-> [scoreA,scoreB]
        # times
        matchestime=matchcard.find_all("span",{"class":"time"})
        
        # fill excell sheet
        wb=load_workbook("./data.xlsx")
        ws=wb["Sheet1"]
        
        # headers
        ws["A1"].value="Champion"
        ws["B1"].value="TeamA"
        ws["C1"].value="TeamB"
        ws["D1"].value="Time"
        ws["E1"].value="Resutl"
        ws["F1"].value="Match Importance"
        rows=ws.max_row+1
        for i in range(rows,len(matchstate)+rows):
            
            
            #teama
            teama=teamA[i-rows].text.strip()
            teamb=teamB[i-rows].text.strip()
            #time
            time=matchestime[i-rows].text.strip()
            #result
            result=f"{scoresA[i-rows].text.strip()}-{scoresB[i-rows].text.strip()}"
            #importance
            state=matchstate[i-rows].text.strip()
            ws[f"A{i}"].value=champion_title
            ws[f"B{i}"].value=teama
            ws[f"C{i}"].value=teamb
            ws[f"D{i}"].value=time
            ws[f"E{i}"].value=result
            ws[f"F{i}"].value=state

        wb.save("./data.xlsx")
        wb.close()
        scoresA.clear()
        scoresB.clear()
        


    


main(page)